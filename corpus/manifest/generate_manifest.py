#!/usr/bin/env python3
"""
generate_manifest.py — validate the SafetyLM corpus manifest CSV and (re)build the XLSX.

WHAT THIS DOES
--------------
1. Reads ``corpus_manifest.csv`` (the canonical, machine-readable catalogue — one row per
   source document).
2. VALIDATES it against the SafetyLM corpus schema (see ``docs/03-corpus-strategy.md`` and the
   Phase 1 manifest prompt):
       - the exact 23-column header, in order
       - every required field populated
       - every controlled field holds a valid enum value
       - every ``source_id`` is unique and well-formed ([JURISDICTION]-[CATEGORY]-[SEQUENCE])
       - project invariants (these are NON-NEGOTIABLE — a safety-critical domain):
             * WA  -> pre_harmonisation == "false"   (WA HAS been harmonised since 31 Mar 2022)
             * VIC -> pre_harmonisation == "true"    (VIC is the sole non-harmonised jurisdiction)
             * FED / NZ / AIHS -> pre_harmonisation == "n/a"
             * NT / AIHS rows and any ``standard`` row -> ingestable == "false"
3. Emits ``corpus_manifest.xlsx`` — a formatted, human-readable copy with:
       - row 1 frozen (freeze panes)
       - an autofilter covering every column
       - auto-fitted column widths
       - a bold header row
4. Prints a coverage summary (per-jurisdiction / per-type counts, plus how many rows still
   need human follow-up: unverified URLs, VERIFY-currency, non-ingestable).

If validation fails, the script prints every problem and exits non-zero WITHOUT writing the
XLSX — so a broken manifest can never silently produce a clean-looking spreadsheet.

INPUT  : corpus_manifest.csv   (same directory as this script)
OUTPUT : corpus_manifest.xlsx  (same directory as this script)

DEPENDENCIES: Python standard library (csv) + openpyxl ONLY. Install openpyxl with
``pip install openpyxl`` if it is missing.

USAGE: python3 generate_manifest.py            # validate + build, from anywhere
       python3 generate_manifest.py --check     # validate only, do not write the XLSX
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Schema definition — the single source of truth for what a valid manifest is.
# ---------------------------------------------------------------------------

# Exact column order. The CSV header must match this list precisely.
HEADER = [
    "source_id", "jurisdiction", "regulator", "document_type", "document_title",
    "instrument_id", "url", "url_verified", "verification_method", "file_format",
    "date_published", "commencement_date", "currency_version_date", "legislative_currency",
    "hazard_category", "industry_relevance", "priority_tier", "pre_harmonisation",
    "license_notes", "ingestable", "ingest_caveat", "download_attempted", "notes",
]

# Fields that must never be empty. (instrument_id, the three date fields, ingest_caveat,
# download_attempted and notes are allowed to be blank.)
REQUIRED_NONEMPTY = [
    "source_id", "jurisdiction", "regulator", "document_type", "document_title", "url",
    "url_verified", "verification_method", "file_format", "legislative_currency",
    "hazard_category", "industry_relevance", "priority_tier", "pre_harmonisation",
    "license_notes", "ingestable",
]

# Controlled-vocabulary (enum) fields.
JURISDICTIONS = {"FED", "NSW", "VIC", "QLD", "SA", "WA", "TAS", "NT", "ACT", "CTH", "NZ", "AIHS"}
DOC_TYPES = {
    "act", "regulation", "code_of_practice", "compliance_code", "acop",
    "good_practice_guideline", "guidance", "statistical_report", "bok_chapter",
    "standard", "court_decision", "policy",
    # added 2026-06-30 per gap review (new in-scope classes):
    "bill", "safe_work_instrument", "safety_alert", "research_report",
}
CURRENCY = {"CURRENT", "SUPERSEDED", "UNDER_REVIEW", "VERIFY"}
VERIFICATION_METHODS = {"direct_fetch", "search_snippet", "regulator_index"}
FILE_FORMATS = {"PDF", "HTML", "DOCX", "XLSX"}
HAZARDS = {
    "PSYCHOSOCIAL", "OVA", "FALLS", "PLANT_EQUIPMENT", "CHEMICAL", "MANUAL_HANDLING",
    "ELECTRICAL", "CONFINED_SPACES", "CONTRACTOR", "GENERAL",
    # added 2026-06-30 per gap review (eliminates NOISE split + physical-agent mis-tags):
    "NOISE", "BIOLOGICAL", "RADIATION", "THERMAL",
}
INDUSTRIES = {"CONSTRUCTION", "HEALTH", "TRANSPORT", "MANUFACTURING", "MINING", "ALL"}
TIERS = {"1", "2", "3"}
PRE_HARM = {"true", "false", "n/a"}
BOOLS = {"true", "false"}

# Map document_type -> the [CATEGORY] token used inside source_id.
CATEGORY = {
    "act": "LEG", "regulation": "REG", "code_of_practice": "COP", "compliance_code": "CC",
    "acop": "ACOP", "good_practice_guideline": "GPG", "guidance": "GUI",
    "statistical_report": "STAT", "bok_chapter": "BOK", "standard": "STD",
    "court_decision": "CASE", "policy": "POL",
    "bill": "BILL", "safe_work_instrument": "SWI", "safety_alert": "ALERT",
    "research_report": "RESEARCH",
}

# Required pre_harmonisation value per jurisdiction (deterministic project invariant).
PRE_HARM_BY_JUR = {
    "FED": "n/a", "NZ": "n/a", "AIHS": "n/a", "VIC": "true", "WA": "false",
    "NSW": "false", "QLD": "false", "SA": "false", "TAS": "false", "NT": "false",
    "ACT": "false", "CTH": "false",
}

# Recommended licence vocabulary. Values outside this set are WARNINGS, not errors —
# license text legitimately varies, but drift is worth surfacing.
LICENSE_VOCAB = {
    "CC-BY-4.0", "CC-BY-4.0 (VERIFY)", "CC-BY-NC-3.0-AU", "crown-copyright-open-attribution",
    "crown-copyright-restricted-NT", "nz-legislation-public-domain-s27",
    "proprietary-restricted-AIHS", "iso-restricted", "VERIFY",
}

HERE = Path(__file__).resolve().parent
CSV_PATH = HERE / "corpus_manifest.csv"
XLSX_PATH = HERE / "corpus_manifest.xlsx"


def validate(rows: list[dict]) -> list[str]:
    """Return a list of error strings. Empty list => the manifest is valid."""
    errors: list[str] = []
    warnings: list[str] = []
    seen_ids: dict[str, int] = {}

    if not rows:
        return ["CSV contains a header but zero data rows."]

    for i, row in enumerate(rows, start=2):  # start=2: row 1 is the header in the file
        rid = row.get("source_id", "") or f"<row {i}>"

        # 1. required, non-empty
        for field in REQUIRED_NONEMPTY:
            if not (row.get(field) or "").strip():
                errors.append(f"{rid}: required field '{field}' is empty.")

        # 2. enum checks
        jur = (row.get("jurisdiction") or "").strip()
        if jur not in JURISDICTIONS:
            errors.append(f"{rid}: jurisdiction '{jur}' not in {sorted(JURISDICTIONS)}.")
        if (row.get("document_type") or "").strip() not in DOC_TYPES:
            errors.append(f"{rid}: document_type '{row.get('document_type')}' invalid.")
        if (row.get("legislative_currency") or "").strip() not in CURRENCY:
            errors.append(f"{rid}: legislative_currency '{row.get('legislative_currency')}' invalid.")
        if (row.get("verification_method") or "").strip() not in VERIFICATION_METHODS:
            errors.append(f"{rid}: verification_method '{row.get('verification_method')}' invalid.")
        if (row.get("file_format") or "").strip() not in FILE_FORMATS:
            errors.append(f"{rid}: file_format '{row.get('file_format')}' invalid.")
        if (row.get("hazard_category") or "").strip() not in HAZARDS:
            errors.append(f"{rid}: hazard_category '{row.get('hazard_category')}' invalid.")
        if (row.get("priority_tier") or "").strip() not in TIERS:
            errors.append(f"{rid}: priority_tier '{row.get('priority_tier')}' invalid.")
        if (row.get("url_verified") or "").strip() not in BOOLS:
            errors.append(f"{rid}: url_verified '{row.get('url_verified')}' must be true/false.")
        if (row.get("ingestable") or "").strip() not in BOOLS:
            errors.append(f"{rid}: ingestable '{row.get('ingestable')}' must be true/false.")
        if (row.get("pre_harmonisation") or "").strip() not in PRE_HARM:
            errors.append(f"{rid}: pre_harmonisation '{row.get('pre_harmonisation')}' invalid.")

        # industry_relevance is comma-separated; every token must be valid
        ind = (row.get("industry_relevance") or "").strip()
        if ind:
            for tok in [t.strip() for t in ind.split(",") if t.strip()]:
                if tok not in INDUSTRIES:
                    errors.append(f"{rid}: industry_relevance token '{tok}' invalid.")

        # 3. url must look like a URL, never a bare placeholder
        url = (row.get("url") or "").strip()
        if url and not (url.startswith("http://") or url.startswith("https://")):
            errors.append(f"{rid}: url '{url}' is not an http(s) URL (no fabricated/placeholder URLs).")
        # hard rule: only a direct fetch (HTTP 200) may claim url_verified=true; anything
        # confirmed via snippet/index must stay url_verified=false for human follow-up.
        if (row.get("url_verified") or "").strip() == "true" and (row.get("verification_method") or "").strip() != "direct_fetch":
            errors.append(f"{rid}: url_verified='true' requires verification_method='direct_fetch'.")

        # 4. source_id uniqueness + shape [JUR]-[CAT]-[SEQ]
        sid = (row.get("source_id") or "").strip()
        if sid:
            if sid in seen_ids:
                errors.append(f"{rid}: duplicate source_id '{sid}' (also row {seen_ids[sid]}).")
            else:
                seen_ids[sid] = i
            parts = sid.split("-")
            if len(parts) != 3 or parts[0] not in JURISDICTIONS or not parts[2].isdigit():
                errors.append(f"{rid}: source_id '{sid}' is not [JURISDICTION]-[CATEGORY]-[SEQUENCE].")
            else:
                expected_cat = CATEGORY.get((row.get("document_type") or "").strip())
                if expected_cat and parts[1] != expected_cat:
                    errors.append(
                        f"{rid}: source_id category '{parts[1]}' does not match document_type "
                        f"'{row.get('document_type')}' (expected '{expected_cat}')."
                    )

        # 5. INVARIANTS (hard) — the project's non-negotiables
        if jur in PRE_HARM_BY_JUR:
            expected = PRE_HARM_BY_JUR[jur]
            if (row.get("pre_harmonisation") or "").strip() != expected:
                errors.append(
                    f"{rid}: {jur} must have pre_harmonisation='{expected}', "
                    f"got '{row.get('pre_harmonisation')}'."
                )
        if jur in {"NT", "AIHS"} and (row.get("ingestable") or "").strip() != "false":
            errors.append(f"{rid}: {jur} rows must be ingestable='false' (restricted licence).")
        if (row.get("document_type") or "").strip() == "standard" and (row.get("ingestable") or "").strip() != "false":
            errors.append(f"{rid}: 'standard' rows must be ingestable='false' (purchase-only).")
        # ingestable=false must carry a caveat explaining why
        if (row.get("ingestable") or "").strip() == "false" and not (row.get("ingest_caveat") or "").strip():
            errors.append(f"{rid}: ingestable='false' but ingest_caveat is empty.")

        # 6. licence vocabulary (warning only)
        lic = (row.get("license_notes") or "").strip()
        if lic and lic not in LICENSE_VOCAB:
            warnings.append(f"{rid}: license_notes '{lic}' outside the recommended vocabulary.")

    if warnings:
        print(f"\n  {len(warnings)} warning(s) (non-fatal):")
        for w in warnings[:25]:
            print(f"    - {w}")
        if len(warnings) > 25:
            print(f"    ... and {len(warnings) - 25} more.")

    return errors


def build_xlsx(rows: list[dict]) -> None:
    """Write the formatted XLSX: frozen header, autofilter, auto-fitted widths."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "corpus_manifest"

    # header row
    ws.append(HEADER)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    # data rows, in HEADER order
    for row in rows:
        ws.append([row.get(col, "") for col in HEADER])

    # freeze the header row (everything above A2 stays put while scrolling)
    ws.freeze_panes = "A2"

    # autofilter across the whole used range
    last_col = get_column_letter(len(HEADER))
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # auto-fit column widths: widest cell in the column, capped so prose columns stay sane
    CAP = {"url": 70, "notes": 70, "ingest_caveat": 50, "document_title": 60}
    for idx, col in enumerate(HEADER, start=1):
        longest = len(col)
        for row in rows:
            longest = max(longest, len(str(row.get(col, ""))))
        width = min(longest + 2, CAP.get(col, 45))
        ws.column_dimensions[get_column_letter(idx)].width = max(width, 10)

    wb.save(XLSX_PATH)


def summarise(rows: list[dict]) -> None:
    """Print a coverage + follow-up summary."""
    per_jur: dict[str, int] = {}
    per_type: dict[str, int] = {}
    unverified = sum(1 for r in rows if (r.get("url_verified") or "").strip() != "true")
    needs_currency = sum(1 for r in rows if (r.get("legislative_currency") or "").strip() == "VERIFY")
    non_ingestable = sum(1 for r in rows if (r.get("ingestable") or "").strip() == "false")
    for r in rows:
        per_jur[r["jurisdiction"]] = per_jur.get(r["jurisdiction"], 0) + 1
        per_type[r["document_type"]] = per_type.get(r["document_type"], 0) + 1

    print(f"\n  Total documents: {len(rows)}")
    print("  By jurisdiction: " + ", ".join(f"{k}={v}" for k, v in sorted(per_jur.items())))
    print("  By type:         " + ", ".join(f"{k}={v}" for k, v in sorted(per_type.items())))
    print(f"  Follow-up needed: {unverified} URL(s) unverified, "
          f"{needs_currency} currency=VERIFY, {non_ingestable} not ingestable.")


def main() -> int:
    check_only = "--check" in sys.argv[1:]

    if not CSV_PATH.exists():
        print(f"ERROR: {CSV_PATH} not found.", file=sys.stderr)
        return 1

    with CSV_PATH.open(newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration:
            print("ERROR: CSV is empty.", file=sys.stderr)
            return 1
        if header != HEADER:
            print("ERROR: CSV header does not match the expected schema.", file=sys.stderr)
            print(f"  expected: {HEADER}", file=sys.stderr)
            print(f"  found:    {header}", file=sys.stderr)
            return 1
        rows = [dict(zip(HEADER, rec)) for rec in reader if any(c.strip() for c in rec)]

    errors = validate(rows)
    if errors:
        print(f"\nVALIDATION FAILED — {len(errors)} error(s):", file=sys.stderr)
        for e in errors[:80]:
            print(f"  - {e}", file=sys.stderr)
        if len(errors) > 80:
            print(f"  ... and {len(errors) - 80} more.", file=sys.stderr)
        return 1

    print(f"VALIDATION PASSED — {len(rows)} rows, schema + invariants OK.")
    summarise(rows)

    if check_only:
        print("\n(--check) Skipped XLSX build.")
        return 0

    try:
        build_xlsx(rows)
    except ModuleNotFoundError:
        print("\nERROR: openpyxl is not installed. Run: pip install openpyxl", file=sys.stderr)
        return 1
    print(f"\nWrote {XLSX_PATH.name} (freeze row 1, autofilter, auto-fitted widths).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
